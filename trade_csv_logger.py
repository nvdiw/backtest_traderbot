import pandas as pd
import os
from collections import Counter


class TradeCSVLogger:
    """Lightweight CSV logger.
    - In normal mode it collects rows and writes a CSV on save_csv().
    - In optimize mode (optimize=True) it becomes a no-op to avoid disk I/O
      and reduce per-trade overhead (much faster for grid search).
    """
    COLUMNS = [
        "trade_id",
        "type",
        "open_time",
        "close_time",
        "entry_price",
        "close_price",
        "tactical_balance",
        "balance_before",
        "balance_after",
        "total_assets",
        "amount",
        "profit",
        "profit_percent",
        "pnl_percent",
        "fee_paid",
        "leverage",
        "trade_amount_percent",
        "duration_minutes_total",
        "duration_days",
        "duration_hours",
        "duration_minutes",
        "save_money",
        "profit_percent_per_month",
        "other_open_positions_at_close",
        "reason",
    ]

    def __init__(self, optimize: bool = False):
        self.optimize = bool(optimize)
        if self.optimize:
            # keep only a tiny counter to preserve minimal bookkeeping
            self._count = 0
        else:
            self.rows = []

    def log_trade(
        self,
        trade_id,
        trade_type,
        open_time,
        close_time,
        entry_price,
        close_price,
        tactical_balance,
        total_assets,
        balance_before,
        balance_after,
        margin,
        leverage,
        trade_amount_percent,
        profit,
        profit_percent,
        pnl_percent,
        fee,
        days,
        hours,
        minutes,
        save_money,
        profit_percent_per_month,
        other_open_positions_at_close,
        reason
    ):
        if self.optimize:
            # no per-trade allocations during optimization
            self._count += 1
            return

        self.rows.append({
            "trade_id": trade_id,
            "type": trade_type,
            "open_time": open_time,
            "close_time": close_time,
            "entry_price": entry_price,
            "close_price": close_price,
            "tactical_balance": tactical_balance,
            "total_assets": total_assets,
            "balance_before": balance_before,
            "balance_after": balance_after,
            "amount": margin,
            "leverage": leverage,
            "trade_amount_percent": trade_amount_percent,
            "profit": profit,
            "profit_percent": profit_percent,
            "pnl_percent": pnl_percent,
            "fee_paid": fee,
            "duration_minutes_total": days * 24 * 60 + hours * 60 + minutes,
            "duration_days": days,
            "duration_hours": hours,
            "duration_minutes": minutes,
            "save_money": save_money,
            "profit_percent_per_month": profit_percent_per_month,
            "other_open_positions_at_close": bool(other_open_positions_at_close),
            "reason": reason,
        })

    def save_csv(
        self,
        first_balance,
        final_balance,
        total_profit,
        total_profit_percent,
        total_fee,
        start_time,
        end_time,
        days,
        hours,
        minutes,
        file_name: str = os.path.join("outputs", "trades", "data_orders.csv")
    ):
        if self.optimize:
            # do not write any files during optimization
            return {"rows_logged": getattr(self, "_count", 0)}

        df = pd.DataFrame(self.rows, columns=self.COLUMNS)

        summary_row = {
            "trade_id": None,
            "type": "SUMMARY",
            "open_time": start_time,
            "close_time": end_time,
            "entry_price": None,
            "close_price": None,
            "total_assets": final_balance,
            "balance_before": first_balance,
            "balance_after": final_balance,
            "profit": total_profit,
            "profit_percent": total_profit_percent,
            "fee_paid": total_fee,
            "duration_days": days,
            "duration_hours": hours,
            "duration_minutes": minutes
        }
        summary_row_full = {col: summary_row.get(col, None) for col in self.COLUMNS}

        if df.empty:
            df = pd.DataFrame([summary_row_full], columns=self.COLUMNS)
        else:
            df.loc[len(df), self.COLUMNS] = [summary_row_full[col] for col in self.COLUMNS]
        while True:
            try:
                output_dir = os.path.dirname(file_name)
                if output_dir:
                    os.makedirs(output_dir, exist_ok=True)
                df.to_csv(file_name, index=False, encoding="utf-8")
                self._save_colored_excel(df, file_name)
                break
            except PermissionError:
                answer = input(f"please close: {file_name} after close write ok: ")
                if answer == "ok":
                    print("thanks!")

    def _save_colored_excel(self, df: pd.DataFrame, csv_file_name: str):
        """
        CSV cannot store background colors.
        Create a companion XLSX with very light-blue rows for trades
        that share the same close_time (2+ trades closed together).
        """
        if df.empty:
            return

        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
        except Exception:
            return

        excel_file_name = os.path.splitext(csv_file_name)[0] + ".xlsx"
        df.to_excel(excel_file_name, index=False)

        close_times = [str(x) for x in df.get("close_time", [])]
        types = [str(x) for x in df.get("type", [])]
        valid_close_times = [
            ct for ct, t in zip(close_times, types)
            if ct and ct.lower() != "none" and t != "SUMMARY"
        ]
        close_counts = Counter(valid_close_times)
        multi_close_times = {ct for ct, c in close_counts.items() if c >= 2}
        if not multi_close_times:
            return

        wb = load_workbook(excel_file_name)
        ws = wb.active
        fill = PatternFill(fill_type="solid", fgColor="EAF4FF")  # very light blue

        # Row 1 is header, data starts from row 2
        for idx, row in enumerate(df.itertuples(index=False), start=2):
            row_type = str(getattr(row, "type", ""))
            row_close_time = str(getattr(row, "close_time", ""))
            if row_type == "SUMMARY":
                continue
            if row_close_time in multi_close_times:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=idx, column=col).fill = fill

        wb.save(excel_file_name)
